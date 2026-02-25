# main.py - ПОЛНАЯ ИСПРАВЛЕННАЯ ВЕРСИЯ С ПОДДЕРЖКОЙ МНОЖЕСТВЕННЫХ ИЗОБРАЖЕНИЙ
from fastapi import FastAPI, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import create_engine, Column, Integer, String, Text, Float, DateTime, Boolean, ForeignKey, inspect, text, MetaData
from sqlalchemy.orm import declarative_base, sessionmaker, Session, relationship
from datetime import datetime
import os
import shutil
import hashlib
import json
from typing import List

# Создаем папки
os.makedirs("static/uploads/products", exist_ok=True)
os.makedirs("static/uploads/categories", exist_ok=True)
os.makedirs("static/uploads/team", exist_ok=True)
os.makedirs("templates/admin", exist_ok=True)
os.makedirs("static/css", exist_ok=True)
os.makedirs("static/js", exist_ok=True)

app = FastAPI(title="Королевские Жалюзи - Премиум производство")

# Статические файлы
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ===== ДОБАВЛЯЕМ ФИЛЬТР from_json ДЛЯ JINJA2 =====
def from_json(value):
    """Преобразует JSON строку в объект Python"""
    if not value:
        return []
    try:
        return json.loads(value)
    except:
        return []

# Регистрируем фильтр в Jinja2
templates.env.filters['from_json'] = from_json

# База данных
SQLALCHEMY_DATABASE_URL = "sqlite:///./royal_blinds.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Очищаем существующие метаданные и создаем новые
metadata = MetaData()
Base = declarative_base(metadata=metadata)

# ===== МОДЕЛИ =====

class Admin(Base):
    __tablename__ = "admins"
    __table_args__ = {'extend_existing': True}
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    password_hash = Column(String)
    created_at = Column(DateTime, default=datetime.now)

class Category(Base):
    __tablename__ = "categories"
    __table_args__ = {'extend_existing': True}
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True)
    description = Column(Text)
    image = Column(String, nullable=True)
    created_at = Column(DateTime, default=datetime.now)
    products = relationship("Product", back_populates="category")

class Product(Base):
    __tablename__ = "products"
    __table_args__ = {'extend_existing': True}
    id = Column(Integer, primary_key=True, index=True)
    product_id = Column(String, unique=True, index=True)
    name = Column(String)
    category_id = Column(Integer, ForeignKey("categories.id"))
    description = Column(Text)
    price = Column(Float, nullable=True)
    image = Column(String, nullable=True)  # Для обратной совместимости (первое фото)
    images = Column(Text, nullable=True)   # JSON массив путей ко всем фото
    material = Column(String, nullable=True)
    sizes = Column(String, nullable=True)
    in_stock = Column(Boolean, default=True)
    is_popular = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.now)
    category = relationship("Category", back_populates="products")
    
    def get_images_list(self):
        """Возвращает список всех изображений"""
        if self.images:
            try:
                return json.loads(self.images)
            except:
                pass
        # Если есть только старое поле image, возвращаем его как список
        if self.image:
            return [self.image]
        return []
    
    def get_first_image(self):
        """Возвращает первое изображение для совместимости"""
        images = self.get_images_list()
        return images[0] if images else None

class Installer(Base):
    __tablename__ = "installers"
    __table_args__ = {'extend_existing': True}
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    username = Column(String, unique=True, index=True, nullable=False)
    phone = Column(String, nullable=False)
    address = Column(String, nullable=True)
    password_hash = Column(String, nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.now)
    orders = relationship("Order", back_populates="installer")

class Order(Base):
    __tablename__ = "orders"
    __table_args__ = {'extend_existing': True}
    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(String, unique=True, index=True)
    client_name = Column(String, nullable=False)
    client_phone = Column(String, nullable=False)
    
    # Категории
    category_plisse = Column(Boolean, default=False)
    category_daynight = Column(Boolean, default=False)
    category_mini = Column(Boolean, default=False)
    
    # JSON данные для множественных позиций
    plisse_items = Column(Text, nullable=True, default="[]")  # JSON массив позиций Плиссе
    daynight_items = Column(Text, nullable=True, default="[]")  # JSON массив позиций День и Ночь
    mini_items = Column(Text, nullable=True, default="[]")  # JSON массив позиций Мини
    
    # Общие суммы
    total_sum = Column(Float, default=0.0)
    
    # Установщик
    installer_id = Column(Integer, ForeignKey("installers.id"), nullable=True)
    installer_phone = Column(String, nullable=True)
    installer_username = Column(String, nullable=True)
    
    # Статусы
    status = Column(String, default='pending')
    payment_status = Column(String, default='unpaid')
    
    # Даты
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Связи
    installer = relationship("Installer", back_populates="orders")

# ===== СОЗДАНИЕ ТАБЛИЦ =====
Base.metadata.create_all(bind=engine)

# ===== АВТОМАТИЧЕСКОЕ ОБНОВЛЕНИЕ ТАБЛИЦ =====
def upgrade_database():
    """Добавляет новые колонки в существующие таблицы"""
    inspector = inspect(engine)
    
    # Обновление таблицы products
    if inspector.has_table("products"):
        existing_columns = [col['name'] for col in inspector.get_columns("products")]
        
        # Добавляем колонку images если её нет
        if 'images' not in existing_columns:
            try:
                with engine.connect() as conn:
                    conn.execute(text("ALTER TABLE products ADD COLUMN images TEXT"))
                    conn.commit()
                    print("✅ Добавлена колонка: images в таблицу products")
            except Exception as e:
                print(f"⚠️ Не удалось добавить images: {e}")
    
    # Обновление таблицы orders
    if inspector.has_table("orders"):
        existing_columns = [col['name'] for col in inspector.get_columns("orders")]
        
        new_columns = {
            'plisse_items': "ALTER TABLE orders ADD COLUMN plisse_items TEXT DEFAULT '[]'",
            'daynight_items': "ALTER TABLE orders ADD COLUMN daynight_items TEXT DEFAULT '[]'",
            'mini_items': "ALTER TABLE orders ADD COLUMN mini_items TEXT DEFAULT '[]'"
        }
        
        for col_name, sql in new_columns.items():
            if col_name not in existing_columns:
                try:
                    with engine.connect() as conn:
                        conn.execute(text(sql))
                        conn.commit()
                        print(f"✅ Добавлена колонка: {col_name}")
                except Exception as e:
                    print(f"⚠️ Не удалось добавить {col_name}: {e}")

# Запускаем обновление
upgrade_database()

# Проверяем и создаем новые таблицы
def create_new_tables():
    inspector = inspect(engine)
    if not inspector.has_table("installers"):
        Installer.__table__.create(bind=engine)
        print("✅ Таблица 'installers' создана")
    if not inspector.has_table("orders"):
        Order.__table__.create(bind=engine)
        print("✅ Таблица 'orders' создана")

create_new_tables()

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====
def hash_password(password: str):
    return hashlib.sha256(password.encode()).hexdigest()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def init_admin(db: Session):
    admin = db.query(Admin).filter(Admin.username == "admin").first()
    if not admin:
        admin = Admin(
            username="admin",
            password_hash=hash_password("admin123")
        )
        db.add(admin)
        db.commit()
        print("✅ Администратор создан: admin / admin123")

def check_admin_auth(request: Request):
    admin_id = request.cookies.get("admin_id")
    return bool(admin_id)

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ РАБОТЫ С МНОЖЕСТВЕННЫМИ ИЗОБРАЖЕНИЯМИ =====

def save_multiple_images(files: List[UploadFile], product_id: str, prefix: str = "product") -> List[str]:
    """Сохраняет несколько изображений и возвращает список путей"""
    image_paths = []
    
    for index, file in enumerate(files):
        if file and file.filename:
            # Создаем уникальное имя для каждого файла
            ext = file.filename.split(".")[-1]
            timestamp = datetime.now().timestamp()
            filename = f"{prefix}_{product_id}_{index}_{timestamp}.{ext}"
            filepath = f"static/uploads/products/{filename}"
            
            # Сохраняем файл
            with open(filepath, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            # Добавляем путь в список
            image_paths.append(f"/static/uploads/products/{filename}")
            
            # Важно: сбрасываем позицию файла для следующего чтения
            file.file.seek(0)
    
    return image_paths

def delete_image_file(image_path):
    """Удаляет файл изображения"""
    if not image_path:
        return False
    
    try:
        # Получаем абсолютный путь
        if image_path.startswith('/'):
            full_path = image_path[1:]
        else:
            full_path = image_path
        
        # Добавляем текущую директорию если нужно
        if not os.path.isabs(full_path):
            full_path = os.path.join(os.getcwd(), full_path)
        
        if os.path.exists(full_path):
            os.remove(full_path)
            print(f"✅ Файл удален: {full_path}")
            return True
        else:
            print(f"⚠️ Файл не найден: {full_path}")
            return False
    except Exception as e:
        print(f"❌ Ошибка при удалении {image_path}: {e}")
        return False

def delete_multiple_images(images_json):
    """Удаляет несколько изображений по JSON"""
    if not images_json:
        return
    
    try:
        paths = json.loads(images_json)
        for path in paths:
            delete_image_file(path)
    except Exception as e:
        print(f"❌ Ошибка при удалении изображений: {e}")
        
def migrate_existing_products(db: Session):
    """Переносит существующие изображения в новый формат"""
    products = db.query(Product).all()
    migrated_count = 0
    for product in products:
        if product.image and not product.images:
            # Создаем JSON массив из существующего изображения
            product.images = json.dumps([product.image])
            migrated_count += 1
    if migrated_count > 0:
        db.commit()
        print(f"✅ {migrated_count} существующих товаров обновлены для поддержки множественных изображений")

# ===== СТАТИЧЕСКИЕ ФАЙЛЫ =====
@app.get("/static/js/whatsapp.js")
async def whatsapp_js():
    return Response(
        content="""
        function openWhatsApp(sku = '', name = '', price = '', category = '') {
            const phoneNumber = "992201482424";
            let message = "";
            if (sku && name) {
                message = `Здравствуйте! Меня заинтересовал товар с сайта "Королевские Жалюзи":%0A%0A` +
                         `🏷 *Артикул:* ${sku}%0A` +
                         `🪟 *Модель:* ${name}%0A` +
                         `📁 *Категория:* ${category || 'Жалюзи'}%0A` +
                         `💰 *Цена:* ${price || 'по запросу'}%0A%0A` +
                         `Пожалуйста, уточните наличие и условия заказа. Спасибо!`;
            } else {
                message = `Здравствуйте! Меня интересуют жалюзи с сайта "Королевские Жалюзи". Помогите с выбором?`;
            }
            const whatsappUrl = `https://wa.me/${phoneNumber}?text=${message}`;
            window.open(whatsappUrl, '_blank');
        }
        """,
        media_type="application/javascript"
    )

# ===== АВТОРИЗАЦИЯ =====
@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login(request: Request):
    return templates.TemplateResponse("admin/login.html", {"request": request})

@app.post("/admin/login")
async def admin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    admin = db.query(Admin).filter(Admin.username == username).first()
    if admin and admin.password_hash == hash_password(password):
        response = RedirectResponse(url="/admin", status_code=303)
        response.set_cookie(key="admin_id", value=str(admin.id), httponly=True)
        return response
    return templates.TemplateResponse(
        "admin/login.html", 
        {"request": request, "error": "Неверный логин или пароль"}
    )

@app.get("/admin/logout")
async def admin_logout():
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("admin_id")
    return response

# ===== ПУБЛИЧНЫЕ СТРАНИЦЫ =====
@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    init_admin(db)
    migrate_existing_products(db)  # Миграция существующих товаров
    popular_products = db.query(Product).filter(Product.is_popular == True).limit(6).all()
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "index.html", 
        {
            "request": request,
            "popular_products": popular_products,
            "categories": categories,
            "active_page": "home",
            "is_admin": check_admin_auth(request)
        }
    )

@app.get("/about", response_class=HTMLResponse)
async def about(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(
        "about.html",
        {
            "request": request,
            "active_page": "about",
            "is_admin": check_admin_auth(request)
        }
    )

@app.get("/products", response_class=HTMLResponse)
async def products(
    request: Request,
    category: int = None,
    db: Session = Depends(get_db)
):
    query = db.query(Product)
    if category:
        query = query.filter(Product.category_id == category)
    products = query.all()
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "products.html",
        {
            "request": request,
            "products": products,
            "categories": categories,
            "selected_category": category,
            "active_page": "products",
            "is_admin": check_admin_auth(request)
        }
    )

@app.get("/product/{product_id}", response_class=HTMLResponse)
async def product_detail(request: Request, product_id: str, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.product_id == product_id).first()
    if not product:
        return templates.TemplateResponse(
            "404.html",
            {"request": request, "message": "Товар не найден"},
            status_code=404
        )
    category = db.query(Category).filter(Category.id == product.category_id).first()
    related_products = db.query(Product).filter(
        Product.category_id == product.category_id,
        Product.id != product.id
    ).limit(4).all()
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "product_detail.html",
        {
            "request": request,
            "product": product,
            "category": category,
            "related_products": related_products,
            "categories": categories,
            "is_admin": check_admin_auth(request)
        }
    )

# ===== АДМИН ПАНЕЛЬ =====
@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    products = db.query(Product).all()
    categories = db.query(Category).all()
    stats = {
        "total_products": len(products),
        "total_categories": len(categories),
        "in_stock": len([p for p in products if p.in_stock]),
        "popular": len([p for p in products if p.is_popular])
    }
    return templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "products": products,
            "categories": categories,
            "stats": stats
        }
    )

# ===== УПРАВЛЕНИЕ ТОВАРАМИ =====
@app.get("/admin/products", response_class=HTMLResponse)
async def admin_products(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    products = db.query(Product).all()
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "admin/products.html",
        {
            "request": request,
            "products": products,
            "categories": categories
        }
    )

@app.get("/admin/products/edit/{product_id}", response_class=HTMLResponse)
async def admin_edit_product_page(
    product_id: int, 
    request: Request, 
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return RedirectResponse(url="/admin/products?error=not_found", status_code=303)
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "admin/edit_product.html",
        {
            "request": request,
            "product": product,
            "categories": categories
        }
    )

@app.post("/admin/products/add")
async def admin_add_product(
    request: Request,
    product_id: str = Form(...),
    name: str = Form(...),
    category_id: int = Form(...),
    description: str = Form(...),
    price: float = Form(None),
    material: str = Form(None),
    sizes: str = Form(None),
    in_stock: bool = Form(False),
    is_popular: bool = Form(False),
    images: List[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        # Проверка на дубликат
        existing_product = db.query(Product).filter(Product.product_id == product_id).first()
        if existing_product:
            return RedirectResponse(url="/admin/products?error=duplicate_id", status_code=303)
        
        # Сохраняем множественные изображения
        image_paths = []
        if images:
            valid_images = [img for img in images if img and img.filename]
            if valid_images:
                image_paths = save_multiple_images(valid_images, product_id)
        
        # Первое изображение для поля image (для обратной совместимости)
        first_image = image_paths[0] if image_paths else None
        
        product = Product(
            product_id=product_id,
            name=name,
            category_id=category_id,
            description=description,
            price=price,
            material=material,
            sizes=sizes,
            in_stock=in_stock,
            is_popular=is_popular,
            image=first_image,
            images=json.dumps(image_paths) if image_paths else None
        )
        
        db.add(product)
        db.commit()
        
        return RedirectResponse(url="/admin/products?added=1", status_code=303)
        
    except Exception as e:
        print(f"Error adding product: {e}")
        return RedirectResponse(url="/admin/products?error=add_failed", status_code=303)

# ===== ВАЖНО: ЭТО ЕДИНСТВЕННЫЙ ОБРАБОТЧИК ДЛЯ ОБНОВЛЕНИЯ ТОВАРА =====
@app.post("/admin/products/update/{product_id}")
async def admin_update_product(
    product_id: int,
    request: Request,
    product_id_code: str = Form(...),
    name: str = Form(...),
    category_id: int = Form(...),
    description: str = Form(...),
    price: float = Form(None),
    material: str = Form(None),
    sizes: str = Form(None),
    in_stock: bool = Form(False),
    is_popular: bool = Form(False),
    images: List[UploadFile] = File(None),
    image_mode: str = Form("append"),
    image_order: str = Form("[]"),
    deleted_images: str = Form("[]"),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    print("\n" + "="*60)
    print(f"🔄 ОБНОВЛЕНИЕ ТОВАРА ID: {product_id}")
    print("="*60)
    
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            print("❌ Товар не найден")
            return RedirectResponse(url="/admin/products?error=not_found", status_code=303)
        
        print(f"📦 Товар: {product.name}")
        print(f"📋 Полученные данные:")
        print(f"  - Артикул: {product_id_code}")
        print(f"  - Название: {name}")
        print(f"  - Категория: {category_id}")
        print(f"  - Режим: {image_mode}")
        print(f"  - Удаляемые изображения: {deleted_images}")
        
        # Обновляем основные поля
        product.product_id = product_id_code
        product.name = name
        product.category_id = category_id
        product.description = description
        product.price = price if price else None
        product.material = material
        product.sizes = sizes
        product.in_stock = in_stock
        product.is_popular = is_popular
        
        print("✅ Основные поля обновлены")
        
        # Получаем текущие изображения из БД
        current_images = product.get_images_list()
        print(f"🖼️ Текущие изображения в БД ({len(current_images)}): {current_images}")
        
        # ===== ВАЖНО: ОБРАБОТКА УДАЛЕННЫХ ИЗОБРАЖЕНИЙ =====
        if deleted_images and deleted_images != "[]":
            try:
                deleted_list = json.loads(deleted_images)
                print(f"🗑️ СПИСОК НА УДАЛЕНИЕ ИЗ ФОРМЫ: {deleted_list}")
                
                # Удаляем файлы с диска
                for img_path in deleted_list:
                    if img_path:
                        # Получаем полный путь к файлу
                        full_path = img_path[1:] if img_path.startswith('/') else img_path
                        full_path = os.path.join(os.getcwd(), full_path)
                        
                        print(f"  Попытка удалить файл: {full_path}")
                        if os.path.exists(full_path):
                            os.remove(full_path)
                            print(f"  ✅ Файл удален: {full_path}")
                        else:
                            print(f"  ⚠️ Файл не найден: {full_path}")
                
                # Убираем удаленные изображения из списка
                old_count = len(current_images)
                current_images = [img for img in current_images if img not in deleted_list]
                print(f"✅ Удалено из списка: {old_count - len(current_images)} изображений")
                print(f"📋 Осталось после удаления: {current_images}")
                
            except Exception as e:
                print(f"❌ Ошибка при удалении изображений: {e}")
                import traceback
                traceback.print_exc()
        
        # Обработка новых изображений
        new_image_paths = []
        if images:
            valid_images = [img for img in images if img and img.filename]
            if valid_images:
                print(f"📤 Загрузка {len(valid_images)} новых изображений")
                
                for idx, img in enumerate(valid_images):
                    try:
                        # Генерируем уникальное имя файла
                        ext = img.filename.split(".")[-1]
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"product_{product_id_code}_{timestamp}_{idx}.{ext}"
                        filepath = f"static/uploads/products/{filename}"
                        
                        # Сохраняем файл
                        with open(filepath, "wb") as buffer:
                            shutil.copyfileobj(img.file, buffer)
                        
                        new_image_paths.append(f"/static/uploads/products/{filename}")
                        print(f"  ✅ Сохранено: {filename}")
                        
                    except Exception as e:
                        print(f"  ❌ Ошибка при сохранении {img.filename}: {e}")
        
        # Применяем режим
        print(f"🎯 Режим обработки: {image_mode}")
        if image_mode == "replace":
            print("Режим ЗАМЕНА - удаляем все старые изображения")
            # Удаляем все старые файлы
            for img in current_images:
                full_path = img[1:] if img.startswith('/') else img
                full_path = os.path.join(os.getcwd(), full_path)
                if os.path.exists(full_path):
                    os.remove(full_path)
                    print(f"  ✅ Удален старый файл: {full_path}")
            final_images = new_image_paths
        else:  # append
            print("Режим ДОБАВЛЕНИЕ - сохраняем старые и добавляем новые")
            final_images = current_images + new_image_paths
        
        # Применяем порядок изображений
        if image_order and image_order != "[]":
            try:
                ordered_paths = json.loads(image_order)
                print(f"📋 Запрошенный порядок: {ordered_paths}")
                
                # Создаем новый список с правильным порядком
                new_final = []
                for path in ordered_paths:
                    if path in final_images:
                        new_final.append(path)
                
                # Добавляем те, что не были в ordered_paths
                for path in final_images:
                    if path not in new_final:
                        new_final.append(path)
                
                final_images = new_final
                print(f"✅ Применен порядок: {final_images}")
            except Exception as e:
                print(f"❌ Ошибка при применении порядка: {e}")
        
        # Обновляем поля в БД
        product.images = json.dumps(final_images) if final_images else None
        product.image = final_images[0] if final_images else None
        
        print(f"💾 СОХРАНЕНИЕ В БД:")
        print(f"  - Всего изображений: {len(final_images)}")
        print(f"  - images JSON: {product.images}")
        print(f"  - image (первое): {product.image}")
        
        db.commit()
        print("✅ Изменения сохранены в БД")
        print("="*60 + "\n")
        
        return RedirectResponse(url="/admin/products?updated=1", status_code=303)
        
    except Exception as e:
        print(f"❌ Ошибка при обновлении товара: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/admin/products?error=update_failed", status_code=303)

@app.post("/admin/products/delete/{product_id}")
async def admin_delete_product(
    product_id: int, 
    request: Request, 
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if product:
            # Удаляем все изображения
            delete_multiple_images(product.images)
            # Удаляем старое изображение если есть
            if product.image and not product.images:
                try:
                    os.remove(product.image[1:])
                except:
                    pass
            
            db.delete(product)
            db.commit()
        
        return RedirectResponse(url="/admin/products?deleted=1", status_code=303)
        
    except Exception as e:
        print(f"Error deleting product: {e}")
        return RedirectResponse(url="/admin/products?error=delete_failed", status_code=303)

@app.post("/admin/products/toggle-popular/{product_id}")
async def toggle_popular_product(
    product_id: int, 
    request: Request, 
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    product = db.query(Product).filter(Product.id == product_id).first()
    if product:
        product.is_popular = not product.is_popular
        db.commit()
    return RedirectResponse(url="/admin/products", status_code=303)

# ===== УПРАВЛЕНИЕ КАТЕГОРИЯМИ =====
@app.get("/admin/categories", response_class=HTMLResponse)
async def admin_categories(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    categories = db.query(Category).all()
    return templates.TemplateResponse(
        "admin/categories.html",
        {
            "request": request,
            "categories": categories
        }
    )

@app.post("/admin/categories/add")
async def admin_add_category(
    request: Request,
    name: str = Form(...),
    description: str = Form(None),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        image_path = None
        if image and image.filename:
            ext = image.filename.split(".")[-1]
            filename = f"category_{name}_{datetime.now().timestamp()}.{ext}"
            filepath = f"static/uploads/categories/{filename}"
            with open(filepath, "wb") as buffer:
                shutil.copyfileobj(image.file, buffer)
            image_path = f"/static/uploads/categories/{filename}"
        category = Category(
            name=name,
            description=description,
            image=image_path
        )
        db.add(category)
        db.commit()
        return RedirectResponse(url="/admin/categories?added=1", status_code=303)
    except Exception as e:
        print(f"Error adding category: {e}")
        return RedirectResponse(url="/admin/categories?error=add_failed", status_code=303)

@app.post("/admin/categories/update/{category_id}")
async def admin_update_category(
    category_id: int,
    request: Request,
    name: str = Form(...),
    description: str = Form(None),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        category = db.query(Category).filter(Category.id == category_id).first()
        if category:
            category.name = name
            category.description = description
            if image and image.filename:
                if category.image:
                    try:
                        old_image_path = category.image[1:]
                        if os.path.exists(old_image_path):
                            os.remove(old_image_path)
                    except:
                        pass
                ext = image.filename.split(".")[-1]
                filename = f"category_{name}_{datetime.now().timestamp()}.{ext}"
                filepath = f"static/uploads/categories/{filename}"
                with open(filepath, "wb") as buffer:
                    shutil.copyfileobj(image.file, buffer)
                category.image = f"/static/uploads/categories/{filename}"
            db.commit()
            return RedirectResponse(url="/admin/categories?updated=1", status_code=303)
        else:
            return RedirectResponse(url="/admin/categories?error=not_found", status_code=303)
    except Exception as e:
        print(f"Error updating category: {e}")
        return RedirectResponse(url="/admin/categories?error=update_failed", status_code=303)

@app.post("/admin/categories/delete/{category_id}")
async def admin_delete_category(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        category = db.query(Category).filter(Category.id == category_id).first()
        if category:
            if category.image:
                try:
                    image_path = category.image[1:]
                    if os.path.exists(image_path):
                        os.remove(image_path)
                except:
                    pass
            db.delete(category)
            db.commit()
        return RedirectResponse(url="/admin/categories?deleted=1", status_code=303)
    except Exception as e:
        print(f"Error deleting category: {e}")
        return RedirectResponse(url="/admin/categories?error=delete_failed", status_code=303)

# ===== УПРАВЛЕНИЕ УСТАНОВЩИКАМИ =====
@app.get("/admin/installers", response_class=HTMLResponse)
async def admin_installers(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    installers = db.query(Installer).all()
    return templates.TemplateResponse(
        "admin/installers.html",
        {
            "request": request,
            "installers": installers
        }
    )

@app.post("/admin/installers/add")
async def admin_add_installer(
    request: Request,
    name: str = Form(...),
    username: str = Form(...),
    phone: str = Form(...),
    address: str = Form(None),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        existing = db.query(Installer).filter(Installer.username == username).first()
        if existing:
            return RedirectResponse(url="/admin/installers?error=duplicate_username", status_code=303)
        installer = Installer(
            name=name,
            username=username,
            phone=phone,
            address=address if address else None
        )
        db.add(installer)
        db.commit()
        return RedirectResponse(url="/admin/installers?added=1", status_code=303)
    except Exception as e:
        print(f"Error adding installer: {e}")
        return RedirectResponse(url="/admin/installers?error=add_failed", status_code=303)

@app.post("/admin/installers/update/{installer_id}")
async def admin_update_installer(
    installer_id: int,
    request: Request,
    name: str = Form(...),
    username: str = Form(...),
    phone: str = Form(...),
    address: str = Form(None),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        installer = db.query(Installer).filter(Installer.id == installer_id).first()
        if installer:
            existing = db.query(Installer).filter(
                Installer.username == username,
                Installer.id != installer_id
            ).first()
            if existing:
                return RedirectResponse(url="/admin/installers?error=duplicate_username", status_code=303)
            installer.name = name
            installer.username = username
            installer.phone = phone
            installer.address = address if address else None
            db.commit()
            return RedirectResponse(url="/admin/installers?updated=1", status_code=303)
        else:
            return RedirectResponse(url="/admin/installers?error=not_found", status_code=303)
    except Exception as e:
        print(f"Error updating installer: {e}")
        return RedirectResponse(url="/admin/installers?error=update_failed", status_code=303)

@app.post("/admin/installers/delete/{installer_id}")
async def admin_delete_installer(
    installer_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        installer = db.query(Installer).filter(Installer.id == installer_id).first()
        if installer:
            if installer.orders:
                return RedirectResponse(url="/admin/installers?error=has_orders", status_code=303)
            db.delete(installer)
            db.commit()
        return RedirectResponse(url="/admin/installers?deleted=1", status_code=303)
    except Exception as e:
        print(f"Error deleting installer: {e}")
        return RedirectResponse(url="/admin/installers?error=delete_failed", status_code=303)

@app.get("/admin/installers/{installer_id}/orders")
async def get_installer_orders(
    installer_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    try:
        installer = db.query(Installer).filter(Installer.id == installer_id).first()
        if not installer:
            return JSONResponse({"error": "Installer not found"}, status_code=404)
        orders = []
        for order in installer.orders:
            status_class = {
                "pending": "bg-yellow-500 bg-opacity-20 text-yellow-500",
                "production": "bg-blue-500 bg-opacity-20 text-blue-500",
                "ready": "bg-green-500 bg-opacity-20 text-green-500",
                "installed": "bg-purple-500 bg-opacity-20 text-purple-500",
                "cancelled": "bg-red-500 bg-opacity-20 text-red-500"
            }.get(order.status, "")
            status_text = {
                "pending": "🟡 Оформлен",
                "production": "🔵 На производстве",
                "ready": "🟢 Готов",
                "installed": "✅ Установлен",
                "cancelled": "❌ Отменен"
            }.get(order.status, order.status)
            orders.append({
                "id": order.id,
                "order_id": order.order_id,
                "client_name": order.client_name,
                "client_phone": order.client_phone,
                "total_sum": order.total_sum,
                "status": order.status,
                "status_class": status_class,
                "status_text": status_text,
                "payment_status": order.payment_status,
                "created_at": order.created_at.strftime("%d.%m.%Y")
            })
        return JSONResponse({"installer": installer.name, "orders": orders})
    except Exception as e:
        print(f"Error getting installer orders: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

# ===== УПРАВЛЕНИЕ ЗАКАЗАМИ =====

@app.get("/admin/orders", response_class=HTMLResponse)
async def admin_orders(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    installers = db.query(Installer).all()
    return templates.TemplateResponse(
        "admin/orders.html",
        {
            "request": request,
            "orders": orders,
            "installers": installers
        }
    )

@app.post("/admin/orders/add")
async def admin_add_order(
    request: Request,
    client_name: str = Form(""),
    client_phone: str = Form(""),
    category_plisse: bool = Form(False),
    category_daynight: bool = Form(False),
    category_mini: bool = Form(False),
    plisse_items: str = Form("[]"),
    daynight_items: str = Form("[]"),
    mini_items: str = Form("[]"),
    total_sum: float = Form(0.0),
    installer_id: int = Form(None),
    installer_phone: str = Form(None),
    installer_username: str = Form(None),
    status: str = Form("pending"),
    payment_status: str = Form("unpaid"),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        # Генерируем номер заказа
        last_order = db.query(Order).order_by(Order.id.desc()).first()
        if last_order and last_order.order_id:
            try:
                last_num = int(last_order.order_id.split("-")[-1])
                new_num = last_num + 1
            except:
                new_num = 1
        else:
            new_num = 1
        
        order_id = f"ORD-{datetime.now().year}-{new_num:04d}"
        
        # Создаем заказ
        order = Order(
            order_id=order_id,
            client_name=client_name or "Без имени",
            client_phone=client_phone or "Без телефона",
            category_plisse=category_plisse,
            category_daynight=category_daynight,
            category_mini=category_mini,
            plisse_items=plisse_items,
            daynight_items=daynight_items,
            mini_items=mini_items,
            total_sum=total_sum,
            installer_id=installer_id if installer_id else None,
            installer_phone=installer_phone,
            installer_username=installer_username,
            status=status,
            payment_status=payment_status
        )
        
        db.add(order)
        db.commit()
        print(f"✅ Заказ создан: {order_id}, сумма: {total_sum} TJS")
        
        return RedirectResponse(url="/admin/orders?added=1", status_code=303)
        
    except Exception as e:
        print(f"❌ Error adding order: {e}")
        return RedirectResponse(url="/admin/orders?error=add_failed", status_code=303)

@app.post("/admin/orders/update/{order_id}")
async def admin_update_order(
    order_id: int,
    request: Request,
    client_name: str = Form(""),
    client_phone: str = Form(""),
    category_plisse: bool = Form(False),
    category_daynight: bool = Form(False),
    category_mini: bool = Form(False),
    plisse_items: str = Form("[]"),
    daynight_items: str = Form("[]"),
    mini_items: str = Form("[]"),
    total_sum: float = Form(0.0),
    installer_id: int = Form(None),
    installer_phone: str = Form(None),
    installer_username: str = Form(None),
    status: str = Form("pending"),
    payment_status: str = Form("unpaid"),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if order:
            order.client_name = client_name or "Без имени"
            order.client_phone = client_phone or "Без телефона"
            order.category_plisse = category_plisse
            order.category_daynight = category_daynight
            order.category_mini = category_mini
            order.plisse_items = plisse_items
            order.daynight_items = daynight_items
            order.mini_items = mini_items
            order.total_sum = total_sum
            order.installer_id = installer_id if installer_id else None
            order.installer_phone = installer_phone
            order.installer_username = installer_username
            order.status = status
            order.payment_status = payment_status
            order.updated_at = datetime.now()
            
            db.commit()
            print(f"✅ Заказ обновлен: {order.order_id}, сумма: {total_sum} TJS")
            return RedirectResponse(url="/admin/orders?updated=1", status_code=303)
        else:
            return RedirectResponse(url="/admin/orders?error=not_found", status_code=303)
            
    except Exception as e:
        print(f"❌ Error updating order: {e}")
        return RedirectResponse(url="/admin/orders?error=update_failed", status_code=303)

@app.post("/admin/orders/delete/{order_id}")
async def admin_delete_order(
    order_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if order:
            db.delete(order)
            db.commit()
            print(f"✅ Заказ удален: {order.order_id}")
        return RedirectResponse(url="/admin/orders?deleted=1", status_code=303)
    except Exception as e:
        print(f"❌ Error deleting order: {e}")
        return RedirectResponse(url="/admin/orders?error=delete_failed", status_code=303)

# ===== ПОЛУЧЕНИЕ ДАННЫХ ЗАКАЗА ДЛЯ РЕДАКТИРОВАНИЯ =====
@app.get("/admin/orders/get/{order_id}")
async def get_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return JSONResponse({"success": False, "error": "Unauthorized"}, status_code=401)
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return JSONResponse({"success": False, "error": "Order not found"}, status_code=404)
        
        installer = None
        if order.installer_id:
            installer_obj = db.query(Installer).filter(Installer.id == order.installer_id).first()
            if installer_obj:
                installer = {
                    "id": installer_obj.id,
                    "name": installer_obj.name,
                    "username": installer_obj.username,
                    "phone": installer_obj.phone
                }
        
        return JSONResponse({
            "success": True,
            "order": {
                "id": order.id,
                "order_id": order.order_id,
                "client_name": order.client_name or "",
                "client_phone": order.client_phone or "",
                "category_plisse": order.category_plisse,
                "category_daynight": order.category_daynight,
                "category_mini": order.category_mini,
                "plisse_items": json.loads(order.plisse_items) if order.plisse_items else [],
                "daynight_items": json.loads(order.daynight_items) if order.daynight_items else [],
                "mini_items": json.loads(order.mini_items) if order.mini_items else [],
                "total_sum": order.total_sum or 0,
                "installer_id": order.installer_id,
                "installer_phone": order.installer_phone or "",
                "installer_username": order.installer_username or "",
                "installer": installer,
                "status": order.status or "pending",
                "payment_status": order.payment_status or "unpaid",
                "created_at": order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else ""
            }
        })
    except Exception as e:
        print(f"Error getting order: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# ===== ЭКСПОРТ ЗАКАЗОВ В РЕДАКТИРУЕМЫЙ EXCEL =====
from fastapi.responses import FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import tempfile
import json

@app.get("/admin/orders/export-editable-excel")
async def export_editable_excel(request: Request, db: Session = Depends(get_db)):
    if not check_admin_auth(request):
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        # Получаем все заказы
        orders = db.query(Order).order_by(Order.created_at.desc()).all()
        
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file_path = tmp.name
        
        # Создаем рабочую книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        # ===== НАСТРОЙКА ФИЛЬТРОВ =====
        ws.auto_filter.ref = "A1:L1"
        
        # ===== СТИЛИ =====
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_center = Alignment(horizontal="center", vertical="center")
        cell_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Стиль для ячеек с выпадающим списком
        dropdown_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        
        # ===== ЗАГОЛОВКИ =====
        headers = [
            "ID заказа",
            "Дата",
            "Клиент",
            "Телефон клиента",
            "Плиссе (ID | м² | TJS)",
            "День и Ночь (ID | м² | TJS)",
            "Мини (ID | м² | TJS)",
            "Установщик",
            "Телефон установщика",
            "Статус оплаты",
            "Статус заказа",
            "Общая сумма (TJS)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # ===== СОЗДАНИЕ ВЫПАДАЮЩИХ СПИСКОВ =====
        # Для статуса оплаты
        dv_payment = DataValidation(
            type="list", 
            formula1='"✅ Оплачено,⏳ Не оплачено"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Ошибка",
            error="Выберите значение из списка!"
        )
        ws.add_data_validation(dv_payment)
        
        # Для статуса заказа
        dv_status = DataValidation(
            type="list", 
            formula1='"🟡 Оформлен,🔵 На производстве,🟢 Готов,✅ Установлен,❌ Отменен"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Ошибка",
            error="Выберите значение из списка!"
        )
        ws.add_data_validation(dv_status)
        
        # ===== ДАННЫЕ =====
        for row, order in enumerate(orders, 2):
            # Парсим JSON данные
            plisse_items = json.loads(order.plisse_items) if order.plisse_items else []
            daynight_items = json.loads(order.daynight_items) if order.daynight_items else []
            mini_items = json.loads(order.mini_items) if order.mini_items else []
            
            # Форматируем данные для Плиссе
            plisse_text = ""
            for item in plisse_items:
                plisse_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            # Форматируем данные для День и Ночь
            daynight_text = ""
            for item in daynight_items:
                daynight_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            # Форматируем данные для Мини
            mini_text = ""
            for item in mini_items:
                mini_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            # Общая сумма из БД
            total_sum = order.total_sum or 0
            
            # Статусы для выпадающих списков
            payment_status_display = "✅ Оплачено" if order.payment_status == "paid" else "⏳ Не оплачено"
            
            status_display = {
                "pending": "🟡 Оформлен",
                "production": "🔵 На производстве",
                "ready": "🟢 Готов",
                "installed": "✅ Установлен",
                "cancelled": "❌ Отменен"
            }.get(order.status, order.status)
            
            # Данные строки
            row_data = [
                order.order_id,
                order.created_at.strftime("%d.%m.%Y") if order.created_at else "",
                order.client_name,
                order.client_phone,
                plisse_text.strip(),
                daynight_text.strip(),
                mini_text.strip(),
                f"{order.installer.name} (@{order.installer.username})" if order.installer else "Не назначен",
                order.installer_phone or "",
                payment_status_display,
                status_display,
                total_sum
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                if col == 10:
                    dv_payment.add(cell)
                    cell.fill = dropdown_fill
                    cell.alignment = cell_center
                elif col == 11:
                    dv_status.add(cell)
                    cell.fill = dropdown_fill
                    cell.alignment = cell_center
                elif col == 12:
                    cell.alignment = cell_center
                    cell.font = Font(bold=True, color="C5A059")
                    cell.number_format = '#,##0.00 "TJS"'
                elif col in [5, 6, 7]:
                    cell.alignment = cell_left
                else:
                    cell.alignment = cell_left
        
        # ===== АВТОПОДБОР ШИРИНЫ =====
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, len(orders) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    lines = str(cell_value).count('\n') + 1
                    max_line_length = max(len(str(line)) for line in str(cell_value).split('\n'))
                    cell_length = max_line_length * 0.8
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = min(max_length + 2, 50)
            if col in [5, 6, 7]:
                adjusted_width = min(max_length + 5, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        wb.save(file_path)
        
        filename = f"zakazy_edit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"❌ Error exporting editable excel: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

# ===== ОБНОВЛЕНИЕ ЗАКАЗОВ ИЗ EXCEL =====
from fastapi import UploadFile
import pandas as pd
import io

@app.post("/admin/orders/update-from-excel")
async def update_orders_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return JSONResponse({"success": False, "error": "Unauthorized"}, status_code=401)
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name="Заказы")
        
        print(f"📊 Найдено строк в Excel: {len(df)}")
        
        updated_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                order_id_str = str(row.get('ID заказа', '')).strip()
                if pd.isna(order_id_str) or not order_id_str:
                    continue
                
                print(f"🔍 Обработка заказа: {order_id_str}")
                
                order = db.query(Order).filter(Order.order_id == order_id_str).first()
                if not order:
                    errors.append(f"Строка {index + 2}: Заказ {order_id_str} не найден")
                    continue
                
                payment_status = row.get('Статус оплаты')
                if pd.notna(payment_status):
                    old_payment = order.payment_status
                    if "✅" in str(payment_status):
                        order.payment_status = "paid"
                    elif "⏳" in str(payment_status):
                        order.payment_status = "unpaid"
                    print(f"  Статус оплаты: {old_payment} -> {order.payment_status}")
                
                order_status = row.get('Статус заказа')
                if pd.notna(order_status):
                    old_status = order.status
                    status_map = {
                        "🟡 Оформлен": "pending",
                        "🔵 На производстве": "production",
                        "🟢 Готов": "ready",
                        "✅ Установлен": "installed",
                        "❌ Отменен": "cancelled"
                    }
                    status_str = str(order_status).strip()
                    if status_str in status_map:
                        order.status = status_map[status_str]
                        print(f"  Статус заказа: {old_status} -> {order.status}")
                
                client_name = row.get('Клиент')
                if pd.notna(client_name) and str(client_name).strip():
                    old_name = order.client_name
                    order.client_name = str(client_name).strip()
                    print(f"  Имя клиента: {old_name} -> {order.client_name}")
                
                client_phone = row.get('Телефон клиента')
                if pd.notna(client_phone) and str(client_phone).strip():
                    old_phone = order.client_phone
                    order.client_phone = str(client_phone).strip()
                    print(f"  Телефон: {old_phone} -> {order.client_phone}")
                
                total_sum = row.get('Общая сумма (TJS)')
                if pd.notna(total_sum):
                    try:
                        old_sum = order.total_sum
                        if isinstance(total_sum, str):
                            total_sum = float(total_sum.replace('TJS', '').strip())
                        else:
                            total_sum = float(total_sum)
                        order.total_sum = total_sum
                        print(f"  Сумма: {old_sum} -> {order.total_sum}")
                    except Exception as e:
                        print(f"  Ошибка при обновлении суммы: {e}")
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Строка {index + 2}: {str(e)}")
                print(f"❌ Ошибка: {e}")
        
        if updated_count > 0:
            db.commit()
            print(f"✅ Сохранено {updated_count} изменений")
        
        return JSONResponse({
            "success": True,
            "updated": updated_count,
            "errors": errors
        })
        
    except Exception as e:
        print(f"❌ Error updating from excel: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# ===== ОБНОВЛЕНИЕ EXCEL ФАЙЛА БЕЗ СКАЧИВАНИЯ =====
@app.post("/admin/orders/refresh-excel-direct")
async def refresh_excel_direct(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not check_admin_auth(request):
        return JSONResponse({"success": False, "error": "Unauthorized"}, status_code=401)
    
    try:
        orders = db.query(Order).order_by(Order.created_at.desc()).all()
        
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = tmp.name
            content = await file.read()
            tmp.write(content)
        
        wb = openpyxl.load_workbook(temp_path)
        ws = wb["Заказы"]
        
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        
        for row, order in enumerate(orders, 2):
            plisse_items = json.loads(order.plisse_items) if order.plisse_items else []
            daynight_items = json.loads(order.daynight_items) if order.daynight_items else []
            mini_items = json.loads(order.mini_items) if order.mini_items else []
            
            plisse_text = ""
            for item in plisse_items:
                plisse_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            daynight_text = ""
            for item in daynight_items:
                daynight_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            mini_text = ""
            for item in mini_items:
                mini_text += f"{item.get('sku', '')} | {item.get('square', 0)}м² | {item.get('sum', 0)} TJS\n"
            
            total_sum = order.total_sum or 0
            
            payment_status_display = "✅ Оплачено" if order.payment_status == "paid" else "⏳ Не оплачено"
            
            status_display = {
                "pending": "🟡 Оформлен",
                "production": "🔵 На производстве",
                "ready": "🟢 Готов",
                "installed": "✅ Установлен",
                "cancelled": "❌ Отменен"
            }.get(order.status, order.status)
            
            row_data = [
                order.order_id,
                order.created_at.strftime("%d.%m.%Y") if order.created_at else "",
                order.client_name,
                order.client_phone,
                plisse_text.strip(),
                daynight_text.strip(),
                mini_text.strip(),
                f"{order.installer.name} (@{order.installer.username})" if order.installer else "Не назначен",
                order.installer_phone or "",
                payment_status_display,
                status_display,
                total_sum
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 12:
                    cell.font = Font(bold=True, color="C5A059")
                    cell.number_format = '#,##0.00 "TJS"'
        
        wb.save(temp_path)
        
        with open(temp_path, 'rb') as f:
            updated_content = f.read()
        
        os.unlink(temp_path)
        
        original_filename = file.filename
        return Response(
            content=updated_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={original_filename}"
            }
        )
        
    except Exception as e:
        print(f"❌ Error refreshing excel: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port)